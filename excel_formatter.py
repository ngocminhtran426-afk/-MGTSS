import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

def format_excel(file_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Queue"
    
    headers = [
        "ID", "Trạng thái", "Link Video", "Nguồn", "Tiêu đề", 
        "Thư mục Output", "File SRT", "File Mô tả", "Video Thành phẩm", 
        "Thời gian thêm", "Thời gian hoàn thành", "Ghi chú"
    ]
    
    # 1. Styling the headers
    header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # 2. Set Column Widths
    widths = {
        'A': 6,   # ID
        'B': 15,  # Trạng thái
        'C': 40,  # Link Video
        'D': 12,  # Nguồn
        'E': 35,  # Tiêu đề
        'F': 35,  # Thư mục Output
        'G': 35,  # File SRT
        'H': 35,  # File Mô tả
        'I': 35,  # Video Thành phẩm
        'J': 20,  # Thời gian thêm
        'K': 20,  # Thời gian hoàn thành
        'L': 40   # Ghi chú
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # 3. Freeze Top Row
    ws.freeze_panes = "A2"

    # 4. Add Data Validation (Dropdowns) for Status column
    dv_status = DataValidation(type="list", formula1='"Chờ,Đang tải,Đang xử lý,Hoàn thành,Lỗi"', allow_blank=True)
    ws.add_data_validation(dv_status)
    dv_status.add('B2:B1000')

    # Formatting basic cells
    for row in range(2, 51):
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if col == 2: # Status center
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.alignment = Alignment(vertical='center')

    # 5. Thêm hướng dẫn mẫu ở dòng 2
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Chờ")
    ws.cell(row=2, column=3, value="https://youtu.be/...")
    ws.cell(row=2, column=12, value="<- Dán link vào Cột C và chọn Trạng thái là Chờ")
    
    # Highlight example row
    example_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    for col in range(1, len(headers) + 1):
        ws.cell(row=2, column=col).fill = example_fill

    wb.save(file_path)
    print("Done formatting")

format_excel('c:/Users/hauho/Desktop/Tool_Packaged/Tool_Packaged/DanhSach_Video.xlsx')
