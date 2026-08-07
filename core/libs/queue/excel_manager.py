import os
from datetime import datetime
try:
    import openpyxl
except ImportError:
    pass

# Định nghĩa các chỉ mục cột (1-indexed cho Excel)
COL_ID = 1
COL_STATUS = 2
COL_LINK = 3
COL_SOURCE = 4
COL_TITLE = 5
COL_OUTPUT_DIR = 6
COL_SRT = 7
COL_DESC = 8
COL_VIDEO_OUT = 9
COL_TIME_ADDED = 10
COL_TIME_DONE = 11
COL_NOTES = 12

HEADERS = [
    "ID", "Trạng thái", "Link Video", "Nguồn", "Tiêu đề", 
    "Thư mục Output", "File SRT", "File Mô tả", "Video Thành phẩm", 
    "Thời gian thêm", "Thời gian hoàn thành", "Ghi chú"
]

class ExcelQueueManager:
    def __init__(self, excel_path):
        self.excel_path = excel_path
        self._ensure_file_exists()

    def _ensure_file_exists(self):
        """Tạo file Excel nếu chưa có"""
        if not os.path.exists(self.excel_path):
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.worksheet.datavalidation import DataValidation
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Queue"
            
            # 1. Styling the headers
            header_font = Font(name='Segoe UI', size=11, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
            center_align = Alignment(horizontal="center", vertical="center")
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for col_num, header in enumerate(HEADERS, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
                
            # 2. Set Column Widths
            widths = { 'A':6, 'B':15, 'C':40, 'D':12, 'E':35, 'F':35, 'G':35, 'H':35, 'I':35, 'J':20, 'K':20, 'L':40 }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            # 3. Freeze Top Row
            ws.freeze_panes = "A2"

            # 4. Add Data Validation (Dropdowns) for Status column
            dv_status = DataValidation(type="list", formula1='"Chờ,Đang tải,Đang xử lý,Hoàn thành,Lỗi"', allow_blank=True)
            ws.add_data_validation(dv_status)
            dv_status.add('B2:B1000')

            # Formatting basic cells
            for row in range(2, 51):
                for col in range(1, len(HEADERS) + 1):
                    c = ws.cell(row=row, column=col)
                    c.border = thin_border
                    if col == 2:
                        c.alignment = Alignment(horizontal='center', vertical='center')
                    else:
                        c.alignment = Alignment(vertical='center')

            wb.save(self.excel_path)
            print(f"[EXCEL] Đã tạo file Queue mới tại: {self.excel_path}")

    def _load_workbook(self):
        return openpyxl.load_workbook(self.excel_path)

    def get_next_pending_task(self):
        """Lấy dòng đầu tiên có trạng thái 'Chờ'"""
        wb = self._load_workbook()
        ws = wb.active
        
        # Bỏ qua dòng 1 (Header), duyệt từ dòng 2
        for row_idx in range(2, ws.max_row + 1):
            status = ws.cell(row=row_idx, column=COL_STATUS).value
            if status and str(status).strip().lower() == "chờ":
                # Lấy dữ liệu
                return {
                    "row_index": row_idx,
                    "id": ws.cell(row=row_idx, column=COL_ID).value or "",
                    "link": str(ws.cell(row=row_idx, column=COL_LINK).value or "").strip(),
                    "source": str(ws.cell(row=row_idx, column=COL_SOURCE).value or "").strip(),
                    "title": str(ws.cell(row=row_idx, column=COL_TITLE).value or "").strip(),
                    "output_dir": str(ws.cell(row=row_idx, column=COL_OUTPUT_DIR).value or "").strip(),
                }
        return None

    def update_status(self, row_index, status, note=""):
        wb = self._load_workbook()
        ws = wb.active
        ws.cell(row=row_index, column=COL_STATUS).value = status
        
        if note:
            ws.cell(row=row_index, column=COL_NOTES).value = note
            
        if status.lower() == "hoàn thành" or status.lower() == "lỗi":
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws.cell(row=row_index, column=COL_TIME_DONE).value = now
            
        wb.save(self.excel_path)

    def update_outputs(self, row_index, title="", out_dir="", srt="", desc="", video=""):
        wb = self._load_workbook()
        ws = wb.active
        if title: ws.cell(row=row_index, column=COL_TITLE).value = title
        if out_dir: ws.cell(row=row_index, column=COL_OUTPUT_DIR).value = out_dir
        if srt: ws.cell(row=row_index, column=COL_SRT).value = srt
        if desc: ws.cell(row=row_index, column=COL_DESC).value = desc
        if video: ws.cell(row=row_index, column=COL_VIDEO_OUT).value = video
        
        wb.save(self.excel_path)

if __name__ == '__main__':
    import sys, json
    if len(sys.argv) < 3:
        sys.exit(1)
    
    excel_path = sys.argv[1]
    cmd = sys.argv[2]
    
    manager = ExcelQueueManager(excel_path)
    
    if cmd == 'get_next':
        task = manager.get_next_pending_task()
        print(json.dumps(task) if task else 'null')
    elif cmd == 'update_status':
        row = int(sys.argv[3])
        status = sys.argv[4]
        note = sys.argv[5] if len(sys.argv) > 5 else ''
        manager.update_status(row, status, note)
        print('ok')
    elif cmd == 'update_outputs':
        row = int(sys.argv[3])
        video = sys.argv[4]
        manager.update_outputs(row, video=video)
        print('ok')
